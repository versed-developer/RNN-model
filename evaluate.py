import os

import numpy as np
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from tensorflow.keras.models import load_model
from lstm import load_data

model_path = './lstm_model.keras'
phase_list = ['A', 'B', 'C']


def add_or_update_excel_sheet(file_path, sheet_name, data):
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()

    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
        workbook.save(file_path)

    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl') as writer:
        data.to_excel(writer, sheet_name=sheet_name)

    # workbook.save(file_path)


def get_phase_labels(predicted_values):
    rounded_values = np.round(predicted_values).astype(int)
    phase_names = np.array(phase_list)
    phase_indices = np.argmax(rounded_values, axis=1)
    phase_labels = phase_names[phase_indices]
    return phase_labels


def evaluate_saved_models():
    if os.path.exists(model_path):
        model = load_model(model_path)
        print("Model loaded successfully.")

        for name in ['training', 'testing']:
            x_eval, y_eval, index = load_data(f'{name}.xlsx')
            predicted_values = model.predict(x_eval)
            phase_labels = get_phase_labels(predicted_values)
            predicted_df = pd.DataFrame(phase_labels, index=index, columns=['Predicted_PHASE'])
            add_or_update_excel_sheet(f'{name}.xlsx', 'Predicted_Phases', predicted_df)
    else:
        print(f"The file '{model_path}' does not exist.")


if __name__ == '__main__':
    evaluate_saved_models()
