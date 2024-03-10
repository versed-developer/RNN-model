import os
import shutil

import numpy as np
import pandas as pd
import tensorflow.keras as K
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from tensorflow.keras.layers import Dense, LSTM, Concatenate, Input, Attention, Bidirectional
from tensorflow.keras.models import Model
from tensorflow.keras.utils import to_categorical
from tensorflow.keras.callbacks import ModelCheckpoint

phase_list = ['A', 'B', 'C']
num_columns = 0
value_mapping = {0: 'A', 1: 'B', 2: 'C'}


def load_data(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(file_path)

    full_sheet = pd.read_excel(file_path, sheet_name='Full')
    global num_columns
    num_columns = len(full_sheet.columns) - 1

    columns = [f"t{i}" for i in range(1, num_columns + 1)]
    data_full_train = pd.read_excel(file_path, sheet_name='Full', index_col=0, names=columns)
    g_truth_data = pd.read_excel(file_path, sheet_name='G truth')

    node_to_phase = {}
    for phase, nodes in g_truth_data.items():
        for node in nodes.dropna():
            node_to_phase[node] = phase_list.index(phase)

    data_full_train['PHASE'] = data_full_train.index.map(node_to_phase)

    x = np.expand_dims(data_full_train.drop(columns=['PHASE']).to_numpy(), axis=(2,))
    y = to_categorical(data_full_train['PHASE'].to_numpy(), num_classes=3)

    return x, y, data_full_train.index, data_full_train['PHASE']


def build_rnn_model():
    _input = Input(shape=(num_columns, 1))

    lstm_out = Bidirectional(LSTM(128))(_input)
    #lstm_out = Dropout(0.4)(lstm_out)

    attention = Attention()([lstm_out, lstm_out])
    out = Concatenate()([lstm_out, attention])

    out = Dense(512, activation='relu')(out)
    out = Dense(256, activation='relu')(out)
    out = Dense(3, activation='softmax')(out)

    rnn_model = Model(inputs=_input, outputs=out)

    rnn_model.compile(loss='categorical_crossentropy', optimizer='adam', metrics=['accuracy'])
    print(rnn_model.summary())

    return rnn_model


def remove_logs(path):
    for filename in os.listdir(path):
        file_path = os.path.join(path, filename)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print(f"Failed to delete {file_path}. Reason: {e}")


def add_or_update_excel_sheet(file_path, sheet_name, data):
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()
        workbook.save(file_path)

    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
        workbook.save(file_path)

    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl') as writer:
        data.to_excel(writer, sheet_name=sheet_name)


def get_phase_labels(predicted):
    rounded_values = np.round(predicted).astype(int)
    phase_names = np.array(phase_list)
    phase_indices = np.argmax(rounded_values, axis=1)
    phase_labels = phase_names[phase_indices]
    return phase_labels


def write_results(predicted, expected, idx, file_path):
    phase_labels = get_phase_labels(predicted)
    predicted_df = pd.DataFrame(phase_labels, index=idx, columns=['Predicted_PHASE'])

    expected = expected.replace(value_mapping).to_frame()
    merged_df = pd.concat([expected, predicted_df], axis=1)

    add_or_update_excel_sheet(file_path, 'Predicted_Phases', merged_df)


if __name__ == '__main__':
    remove_logs('logs')
    train_file_name = input('Please enter excel file name that includes train data:(training?)').strip() or 'training'
    x_train, y_train, index, g_truth = load_data(f'{train_file_name}.xlsx')
    model = build_rnn_model()

    tensorboard_callback = K.callbacks.TensorBoard(log_dir="./logs")
    model_checkpoint_callback = ModelCheckpoint(filepath='model_checkpoint.keras')
    model.fit(x_train, y_train, epochs=256, batch_size=16, callbacks=[tensorboard_callback, model_checkpoint_callback])

    # If it's fit to your consideration, save model
    save_model = input("Save current model?")
    if save_model == "y" or save_model == "Y":
        model.save('lstm_model.keras')
        predicted_values = model.predict(x_train)
        write_results(predicted_values, g_truth, index, f'{train_file_name}_result.xlsx')
